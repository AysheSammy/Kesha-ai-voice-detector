import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from constants import EXTENSIONS, LETTERS

def insert_image_to_excel(excelName):
    filePath = os.path.dirname(os.path.realpath(__file__))
    excelPath = [os.path.join(filePath, f) for f in os.listdir(filePath) if os.path.isfile(os.path.join(filePath, f)) and f.split('.')[-1] in EXTENSIONS['excel'] and f.split('.')[0] == excelName]
    print(excelPath)
    if excelPath != []:
        wb = load_workbook(excelPath[0])
        sheetNames = wb.sheetnames
        for sheetName in sheetNames:
            print('==================SHEETNAME: ', sheetName, '=================================')
            sheet = wb[sheetName]
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            for c in range(1, max_col+1):
                for r in range(1, max_row+1):
                    print('row: ',r,' col: ', c)
                    cellValue = sheet.cell(row=r, column=c).value
                    if cellValue != None:
                        print(cellValue)

                        if '.img' in cellValue:
                            imgName = os.path.basename(os.path.splitext(cellValue)[0]).split('_')[0]
                            folderPath = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'images')
                            files = os.listdir(folderPath)

                            for file in files:
                                imgFileName = os.path.splitext(file)[0]
                                if imgName == imgFileName:
                                    if os.path.isfile(os.path.join(folderPath, file)):
                                        img = Image(os.path.join(folderPath, file))
                                    else:
                                        imgFolder = os.path.join(folderPath, file)
                                        img = Image(os.path.join(imgFolder, os.listdir(imgFolder)[0]))
                                    cell = f'{LETTERS[c]}{r}'
                                    print('CELL: ', cell)
                                    img.anchor = cell
                                    sheet.add_image(img)
                        else:
                            continue
        wb.save(excelPath[0])
        wb.close()

filename = input("Please enter the excel file name: ")
insert_image_to_excel(filename)