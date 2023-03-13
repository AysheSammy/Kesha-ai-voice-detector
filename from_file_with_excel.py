import os
import time
from openpyxl import load_workbook
from PIL import Image
import pyautogui

from constants import EXTENSIONS, MOUSE_DURATION
from methods import getCoorFromColoredImg, imageFunc, typeText
from tts import _TTS


# functions
def speak(what):
    tts = _TTS()
    tts.start(what)
    del(tts)


def fromFile(excelName):
    filePath = os.path.dirname(os.path.realpath(__file__))
    excelPath = [os.path.join(filePath, f) for f in os.listdir(filePath) if os.path.isfile(os.path.join(filePath, f)) and f.split('.')[-1] in EXTENSIONS['excel'] and f.split('.')[0] == excelName]
    print(excelPath)
    if excelPath != []:
        readExcel(excelPath[0])


def readExcel(excelPath):
    wb = load_workbook(excelPath, read_only=True)
    sheetNames = wb.sheetnames
    for sheetName in sheetNames:
        print('==================SHEETNAME: ', sheetName, '=================================')
        loopThroughSheet(wb[sheetName])
    wb.close()


def loopThroughSheet(sheet, start_row=1, start_col=1):
    """
    Loops through scenarios over the sheet and operates them.
    It's a recursive function, that calls itself when the image is not found.
    And loops through the rows.
    
    If retVal == 0 (If image isn't found)                   => column = column + 1, row = 1
    If retVal == 1 (If image is found or text is written)   => column = column,     row + 1
    
    Args:
        sheet (Excel sheet): Excel sheet
        start_row (int, optional): Starting row number. Defaults to 1.
        start_col (int, optional): Starting column number. Defaults to 1.

    Returns:
        dynamic: calls itself or "Finished"
    """
    max_row = sheet.max_row
    
    for r in range(start_row, max_row+1):
        print('row: ',r,' col: ', start_col)
        cellValue = sheet.cell(row=r, column=start_col).value
        if cellValue != None:
            if cellValue != '-':
                retVal = None
                print(cellValue)

                if '.img' in cellValue:
                    filevars = os.path.splitext(cellValue)
                    fileName = os.path.basename(filevars[0])
                    retVal = operateOverFiles(fileName)
                    print('rertVal ', retVal)
                else:
                    typeText(cellValue)
                    retVal = 1

                if retVal == 0:
                    return loopThroughSheet(sheet, start_row=1, start_col=start_col+1)
                elif retVal == 1:
                    continue
        else:
            return "Finished"


def operateOverFiles(fileName):
    """
    Due to the fileType finds file from the given path and operates it.

    Args:
        fileType (String): File type key from constants
        fileName (String): Name of the file without extension
        isLastCycle (bool): Optional
    """

    filePath = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'images')
    files = os.listdir(filePath)
    pars = fileName.split('_')
    imgName = pars[0]
    actionType = pars[1]
    duration = float(pars[2][1:]) / 10 if float(pars[2][1:]) >= 1 else float(pars[2][1:])
    coordination = tuple(int(r) for r in pars[3][1:].split(' ')) if pars[3][1:] != '' else None

    status = 0
    for file in files:
        imgFileName = os.path.splitext(file)[0]
        if imgName == imgFileName:
            if duration == 8.8:
                pass
                for _ in range(88):
                    time.sleep(0.5)
                    imgFolder = os.path.join(filePath, file)
                    imgFile = ''
                    if os.path.isfile(imgFolder):
                        coor = getCoorFromColoredImg(imgFolder, coordination)
                        imgFile = imgFolder
                    elif os.path.isdir(imgFolder):
                        for f in os.listdir(imgFolder):
                            imgFile = os.path.join(imgFolder, f)
                            tryCoor = getCoorFromColoredImg(imgFile, coordination)
                            if tryCoor != None:
                                coor = tryCoor
                                break
                            else:
                                coor = None

                    if coor != None:
                        status = 1
                        if actionType == '2':
                            img = Image.open(imgFile)
                            width, height = img.width, img.height
                            x, y = coor
                            x = x - (width/2)
                            y = y - (height/2)
                            if coordination == None:
                                coorPar = os.path.splitext(os.path.basename(imgFile))[0].split('-')[-1][1:]
                                coordination = tuple(int(r) for r in coorPar.split(' ')) if coorPar != '' else None
                            x = x + coordination[0] if coordination[0] > 0 else x - coordination[0]
                            y = y + coordination[1] if coordination[1] > 0 else y - coordination[1]
                            coor = tuple([x, y])
                        pyautogui.moveTo(coor, duration=MOUSE_DURATION)
                        pyautogui.click()
                        break
                    else:
                        status = 0
            else:
                time.sleep(duration)
                imgFolder = os.path.join(filePath, file)
                imgFile = ''
                if os.path.isfile(imgFolder):
                    coor = getCoorFromColoredImg(imgFolder, coordination)
                    imgFile = imgFolder
                elif os.path.isdir(imgFolder):
                    for f in os.listdir(imgFolder):
                        imgFile = os.path.join(imgFolder, f)
                        tryCoor = getCoorFromColoredImg(imgFile, coordination)
                        if tryCoor != None:
                            coor = tryCoor
                            break
                        else:
                            coor = None

                if coor != None:
                    status = 1
                    if actionType == '2':
                        img = Image.open(imgFile)
                        width, height = img.width, img.height
                        x, y = coor
                        x = x - (width/2)
                        y = y - (height/2)
                        if coordination == None:
                            coorPar = os.path.splitext(os.path.basename(imgFile))[0].split('-')[-1][1:]
                            coordination = tuple(int(r) for r in coorPar.split(' ')) if coorPar != '' else None
                        x = x + coordination[0] if coordination[0] > 0 else x - coordination[0]
                        y = y + coordination[1] if coordination[1] > 0 else y - coordination[1]
                        coor = tuple([x, y])

                    pyautogui.moveTo(coor, duration=MOUSE_DURATION)
                    pyautogui.click()
                else:
                    status = 0
    return status
